from string import ascii_uppercase
from tkinter import Tk
from tkinter import StringVar
from tkinter import IntVar
from tkinter import Label
from tkinter import Canvas
from tkinter import Button
from tkinter import messagebox
from tkinter import ttk
from tkinter import Checkbutton
from tkinter import Text
from tkinter import PhotoImage
from tkinter import Toplevel
from tkinter import Radiobutton
from tkinter import Entry
import openpyxl
import matplotlib.figure
import matplotlib.patches
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkcalendar import Calendar
from fpdf import FPDF
from datetime import date
from datetime import datetime
from operator import itemgetter
from math import inf
from matplotlib.colors import to_rgba_array
import pandas

from tksheet import Sheet
from pynput import keyboard
from PIL import ImageTk, Image
import json
import threading
import time
import sqlite3
import hashlib

# genRight if includeSearch = False: load data from local json
# genRight if includeSearch = True: don't load data from local json

chungTaDangODau = ''
isSaved = True
cfg_radius = 60
cfg_clr = 'white'
winClosed = False
tblWidth = 825
tblHeight = 425
appWidth = 1200
appHeight = 768
appIco = ''
navWidth = appWidth/4

heightLineNav = appHeight/3
lblNavX = navWidth/5*2+10
tabSel = [True, False, False, False, False, False, False]

# fetching data from local db files

con = sqlite3.connect('example.db')
curGlobal = con.cursor()

curGlobal.execute('''SELECT * FROM phieuThue''')
dataPhieuThue = curGlobal.fetchall()
tmp = []
for row in dataPhieuThue: tmp.append(list(row))
dataPhieuThue = tmp

curGlobal.execute('''SELECT * FROM nhanVien''')
dataNv = curGlobal.fetchall()
tmp = []
for row in dataNv: tmp.append(list(row))
dataNv = tmp

curGlobal.execute('''SELECT * FROM phong''')
dataPh = curGlobal.fetchall()
tmp = []
for row in dataPh: tmp.append(list(row))
dataPh = tmp

curGlobal.execute('''SELECT * FROM khachHang''')
dataKh = curGlobal.fetchall()
tmp = []
for row in dataKh: tmp.append(list(row))
dataKh = tmp

curGlobal.execute('''SELECT * FROM dichVu''')
dataDv = curGlobal.fetchall()
tmp = []
for row in dataDv: tmp.append(list(row))
dataDv = tmp

curGlobal.execute('''SELECT * FROM pNhapTbiAndFood''')
phieuNhapTbAndFood = curGlobal.fetchall()
tmp = []
for row in phieuNhapTbAndFood: tmp.append(list(row))
phieuNhapTbAndFood = tmp

def fetchFromDb():
    curGlobal.execute('''SELECT * FROM phieuThue''')
    global dataPhieuThue
    dataPhieuThue = curGlobal.fetchall()
    tmp = []
    for row in dataPhieuThue: tmp.append(list(row))
    dataPhieuThue = tmp

    curGlobal.execute('''SELECT * FROM nhanVien''')
    global dataNv
    dataNv = curGlobal.fetchall()
    tmp = []
    for row in dataNv: tmp.append(list(row))
    dataNv = tmp

    curGlobal.execute('''SELECT * FROM phong''')
    global dataPh
    dataPh = curGlobal.fetchall()
    tmp = []
    for row in dataPh: tmp.append(list(row))
    dataPh = tmp

    curGlobal.execute('''SELECT * FROM khachHang''')
    global dataKh
    dataKh = curGlobal.fetchall()
    tmp = []
    for row in dataKh: tmp.append(list(row))
    dataKh = tmp

    curGlobal.execute('''SELECT * FROM dichVu''')
    global dataDv
    dataDv = curGlobal.fetchall()
    tmp = []
    for row in dataDv: tmp.append(list(row))
    dataDv = tmp

    curGlobal.execute('''SELECT * FROM pNhapTbiAndFood''')
    global phieuNhapTbAndFood
    phieuNhapTbAndFood = curGlobal.fetchall()
    tmp = []
    for row in phieuNhapTbAndFood: tmp.append(list(row))
    phieuNhapTbAndFood = tmp

    return


# initialize database by lists
# dataPhieuThue = [
#     ['PTP0001', 'KH001', 'M001', 'DV001', '20/10/2021', '25/10/2021', False, 0, 0, 0, 'A101'],
#     ['PTP0002', 'KH002', 'M001', 'DV002', '30/01/2022', '17/02/2022', False, 0, 0, 0, 'A102'],
#     ['PTP0003', 'KH001', 'M002', 'DV003', '18/06/2015', '28/06/2015', True, 0, 0, 0, 'A102'],
#     ['PTP0004', 'KH003', 'M002', 'DV008', '08/01/2016', '17/02/2016', False, 0, 0, 0, 'B202'],
#     ['PTP0005', 'KH004', 'M002', 'DV008', '08/01/2018', '17/02/2018', True, 0, 0, 0, 'B202'],
#     ['PTP0006', 'KH006', 'M001', 'DV004', '08/01/2019', '17/02/2019', False, 0, 0, 0, 'B202'],
# ]
# phieuNhapTbAndFood = [
#     ['P001','H001', 'Coca-cola lon', 1000, 'NCC001', 'Nhà cung cấp 1', 'NV01', 500000, '10/03/2022'],
#     ['P002','H002', 'Pepsi lon', 1000, 'NCC002', 'Nhà cung cấp 2', 'NV01', 500000, '11/03/2022'],
#     ['P003','H003', 'Bia Sài Gòn', 1000, 'NCC003', 'Nhà cung cấp 3', 'NV01', 500000, '12/03/2022'],
#     ['P004','H004', 'Trái cây', 20, 'NCC004', 'Nhà cung cấp 4', 'NV01', 100000, '14/03/2022'],
#     ['P005','H005', 'Khăn lạnh', 1000, 'NCC005', 'Nhà cung cấp 5', 'NV01', 100000, '19/03/2022'],
#     ['P006','H006', 'Thực phẩm', 20, 'NCC006', 'Nhà cung cấp 6', 'NV02', 200000, '21/03/2022'],
#     ['P007','H007', 'Pepsi lon', 1000, 'NCC002', 'Nhà cung cấp 7', 'NV01', 500000, '25/03/2022'],
# ]
# dataDv = [
#     ['DV001', 'Xông hơi', 250000],
#     ['DV002', 'Massage', 500000],
#     ['DV003', 'Tắm hồ bơi', 200000],
#     ['DV004', 'Giặt ủi', 50000],
#     ['DV005', 'Ăn sáng', 500000],
#     ['DV006', 'Ăn trưa', 250000],
#     ['DV007', 'Ăn tối', 250000]
# ]
# dataNv = [
#     ['B001', 'Trịnh Quang', 'Hòa', 'Nam', 1979, 'TP.HCM', 'CEO', 'Hội đồng quản trị', 99999999, 99999999],
#     ['B002', 'Kim Đức', 'Long', 'Nam', 1983, 'Quảng Nam', 'CTO', 'Hội đồng quản trị', 99999999, 99999999],
#     ['B003', 'Huỳnh Nguyên', 'Khang', 'Nam', 1982, 'Bình Phước', 'CFO', 'Hội đồng quản trị', 99999999, 99999999],
#     ['G001', 'Hoàng Hòa', 'Hợp', 'Nam', 1982, 'Bình Thuận', 'Nhân viên', 'Bảo vệ', 5000000, 800000],
#     ['G002', 'Lưu Duy', 'Hiếu', 'Nam', 1991, 'Long An', 'Nhân viên', 'Bảo vệ', 5000000, 800000],
#     ['G003', 'Hồ Văn', 'Thông', 'Nam', 1998, 'Gia Lai', 'Trưởng phòng', 'Bảo vệ', 6000000, 1000000],
#     ['G004', 'Nguyễn Văn', 'Sinh', 'Nam', 1991, 'TP.HCM', 'Nhân viên', 'Bảo vệ', 5000000, 600000],
#     ['E001', 'Châu Văn', 'Đạt', 'Nam', 1998, 'Quảng Nam', 'Trưởng phòng', 'Kế toán', 8000000, 600000],
#     ['E002', 'Nguyễn Thị', 'Nga', 'Nữ', 1991, 'Long An', 'Nhân viên', 'Kế toán', 7000000, 600000],
#     ['E003', 'Vũ Việt', 'Đông', 'Nam', 1982, 'Tiền Giang', 'Nhân viên', 'Kế toán', 7000000, 600000],
#     ['C001', 'Đào Ngọc', 'Cẩm', 'Nữ', 1992, 'Bình Dương', 'Nhân viên', 'Vệ sinh', 4000000, 600000],
#     ['C002', 'Huỳnh Thụy Phương', 'Khánh', 'Nữ', 1996, 'TP.HCM', 'Nhân viên', 'Vệ sinh', 4000000, 600000],
#     ['F001', 'Nguyễn Văn Gia', 'Trí', 'Nam', 1992, 'Tiền Giang', 'Trưởng phòng', 'Quản trị nhân sự', 6000000, 600000],
#     ['F002', 'Lưu Ngọc Hoài ', 'Trinh', 'Nữ', 1992, 'Bình Dương', 'Nhân viên', 'Quản trị nhân sự', 6000000, 600000],
#     ['Z001', 'Đoàn Trần Văn ', 'Kha', 'Nam', 1996, 'Long An', 'Bếp trưởng', 'Bếp ăn', 4000000, 600000],
#     ['Z002', 'David ', 'Joe', 'Nam', 1998, 'TP.HCM', 'Đầu bếp', 'Bếp ăn', 8700000, 900000],
#     ['Z003', 'Phạm Toàn ', 'Thắng', 'Nam', 1998, 'TP.HCM', 'Đầu bếp', 'Bếp ăn', 4000000, 600000],
#     ['M001', 'Phạm Quốc ', 'Khải', 'Nam', 1996, 'Tiền Giang', 'Tiếp tân', 'Lễ tân', 3000000, 600000],
#     ['M002', 'Nguyễn Võ ', 'Lợi', 'Nam', 2000, 'Gia Lai', 'Tiếp tân', 'Lễ tân', 3000000, 600000],
#     ['T001', 'Lưu Bích ', 'Thoa', 'Nữ', 1992, 'Quảng Nam', 'Nhân viên', 'massage', 7000000, 600000],
#     ['T002', 'Trần Thục ', 'Quyên', 'Nữ', 2000, 'Tây Ninh', 'Nhân viên', 'massage', 7000000, 600000],
#     ['K001', 'Hứa Vĩnh ', 'Đức', 'Nam', 1990, 'Bến Tre', 'Nhân viên', 'Kho vận', 7000000, 600000],
#     ['K002', 'Trần Phùng ', 'Thọ', 'Nam', 1998, 'TP.HCM', 'Nhân viên', 'Kho vận', 7000000, 600000]
# ]
# dataPh = [
#     ['A101', 500000, 'available', 'Phòng thường 1 giường','normal', "1 giường+1 tu lanh nho+1 bo ban ghe+1 may lanh+2 den"],
#     ['A102', 800000, 'available', 'Phòng thường 2 giường','normal', "2 giường+1 tu lanh nho+1 bo ban ghe+1 may lanh+2 den"],
#     ['B202', 1000000, 'available', 'Phòng cao cấp 1 giường','elite', "1 giường+1 tu lanh lon+2 bo ban ghe +1 may lanh+4 den+1 bon tam"],
#     ['B201', 1200000, 'available', 'Phòng thường 2 giường','elite', "2 giường+1 tu lanh lon+2 bo ban ghe +1 may lanh+4 den+1 bon tam"],
#     ['C301', 3000000, 'available', 'Phòng thượng hạng 1 giường','vip', "1 giường+1 tu lanh lon+3 bo ban ghe +2 may lanh+6 den+1 bon tam+dien thoai+quay bep+minibar"],
#     ['C303', 3200000, 'occupied', 'Phòng thượng hạng 2 giường','vip',"2 giường+1 tu lanh lon+3 bo ban ghe +2 may lanh+6 den+1 bon tam+dien thoai+quay bep+minibar"],
#     ['A103', 600000, 'maintaining', 'Phòng thườnng 1 giường','normal', "giường+ 1 tu lanh nho+1 bo ban ghe+1 may lanh+2 den"]
# ]
# dataKh = [
#     ['KH001', 'Nguyễn Thế', 'Doanh', 'Nam', 1983, '286 Str. 3/2, Ward 12, Dist', 312509075, 'Việt Nam', '09753650117', 'user01@gmail.com' ],
#     ['KH002', 'Úc Quốc', 'Hải', 'Nam', 1982, ' 95 Nguyen Hong Dao street, Tan Binh District', 312509076, 'Việt Nam', '09753650117', 'user02@gmail.com' ],
#     ['KH003', 'Nguyễn Thiện', 'Ân', 'Nam', 1979, '49 Le Trung Nghia, Ward. 12, Tan Binh District', 312509075, 'Việt Nam', '09753650117', 'user03@gmail.com' ],
#     ['KH004', 'Vương Đăng', 'Đạt', 'Nam', 1982, '128 Tran Quy Cap, Group 4, Ninh Hoa, Khanh Hoa', 312509075, 'Việt Nam', '09753650117', 'user04@gmail.com'],
#     ['KH005', 'Trang Diệu', 'Nương', 'Nữ', 1991, 'Tan Quy Tay Ward, Sa Dec Township', 312509075, 'Việt Nam', '09753650117', 'user05@gmail.com' ],
#     ['KH006', 'Nguyễn Chiêu', 'Dương', 'Nữ', 1998, '659 Xo Viet Nghe Tinh, Binh Thanh District', 312509075, 'Việt Nam', '09753650117', 'user06@gmail.com' ],
#     ['KH007', 'Bùi Thúy', 'Vy', 'Nam', 1991, '39A/3 Kha Van Can Street, Hiep Binh Chanh Ward, Thu Duc District', 312509075, 'Việt Nam', '09753650117', 'user07@gmail.com' ],
# ]

def genNav():
    
    lftNavDiv = Canvas(
        dashbrd,
        width = navWidth,
        height = appHeight,
        background = 'white'
    )
    a = [i for i in range(200, 1200, 60)]
    b = [i-10 for i in range(260, 1200, 60)]
    
    def genBut1(x, y, filename):
        global icoBut1
        rawIcoBut1 = Image.open('./img/'+filename)
        rawIcoBut1 = rawIcoBut1.resize((x, y), Image.Resampling.LANCZOS)
        icoBut1 = ImageTk.PhotoImage(rawIcoBut1)
        global lblIcoBut1
        lblIcoBut1 = Label(dashbrd, image = icoBut1, bg = cfg_clr)
    genBut1(200, 60, 'but.png')
    def mouEnBut1(e):
        lbl1['font'] = ('Chirp', 14, 'bold')
    def mouLeBut1(e):
        lbl1['font'] = ('Chirp', 14)

    def genBut2(x, y):
        global icoBut2
        rawIcoBut2 = Image.open('./img/but.png')
        rawIcoBut2 = rawIcoBut2.resize((x, y), Image.Resampling.LANCZOS)
        icoBut2 = ImageTk.PhotoImage(rawIcoBut2)
        global lblIcoBut2
        lblIcoBut2 = Label(dashbrd, image = icoBut2, bg = cfg_clr)
        return
    genBut2(200, 60)
    def mouEnBut2(e):
        lbl2['font'] = ('Chirp', 14, 'bold')
    def mouLeBut2(e):
        lbl2['font'] = ('Chirp', 14)

    def genBut3(x, y):
        global icoBut3
        rawIcoBut3 = Image.open('./img/but.png')
        rawIcoBut3 = rawIcoBut3.resize((x, y), Image.Resampling.LANCZOS)
        icoBut3 = ImageTk.PhotoImage(rawIcoBut3)
        global lblIcoBut3
        lblIcoBut3 = Label(dashbrd, image = icoBut3, bg = cfg_clr)
        return
    genBut3(200, 60)
    def mouEnBut3(e):
        lbl3['font'] = ('Chirp', 14, 'bold')
    def mouLeBut3(e):
        lbl3['font'] = ('Chirp', 14)

    def genBut4(x, y):
        global icoBut4
        rawIcoBut4 = Image.open('./img/but.png')
        rawIcoBut4 = rawIcoBut4.resize((x, y), Image.Resampling.LANCZOS)
        icoBut4 = ImageTk.PhotoImage(rawIcoBut4)
        global lblIcoBut4
        lblIcoBut4 = Label(dashbrd, image = icoBut4, bg = cfg_clr)
        return
    genBut4(200, 60)
    def mouEnBut4(e):
        lbl4['font'] = ('Chirp', 14, 'bold')
    def mouLeBut4(e):
        lbl4['font'] = ('Chirp', 14)

    def genBut5(x, y):
        global icoBut5
        rawIcoBut5 = Image.open('./img/but.png')
        rawIcoBut5 = rawIcoBut5.resize((x, y), Image.Resampling.LANCZOS)
        icoBut5 = ImageTk.PhotoImage(rawIcoBut5)
        global lblIcoBut5
        lblIcoBut5 = Label(dashbrd, image = icoBut5, bg = cfg_clr)
        return
    genBut5(200, 60)
    def mouEnBut5(e):
        lbl5['font'] = ('Chirp', 14, 'bold')
    def mouLeBut5(e):
        lbl5['font'] = ('Chirp', 14)

    def genBut6(x, y):
        global icoBut6
        rawIcoBut6 = Image.open('./img/but.png')
        rawIcoBut6 = rawIcoBut6.resize((x, y), Image.Resampling.LANCZOS)
        icoBut6 = ImageTk.PhotoImage(rawIcoBut6)
        global lblIcoBut6
        lblIcoBut6 = Label(dashbrd, image = icoBut6, bg = cfg_clr)
        return
    genBut6(200, 60)
    def mouEnBut6(e):
        lbl6['font'] = ('Chirp', 14, 'bold')
    def mouLeBut6(e):
        lbl6['font'] = ('Chirp', 14)

    def genBut7(x, y):
        global icoBut7
        rawIcoBut7 = Image.open('./img/but.png')
        rawIcoBut7 = rawIcoBut7.resize((x, y), Image.Resampling.LANCZOS)
        icoBut7 = ImageTk.PhotoImage(rawIcoBut7)
        global lblIcoBut7
        lblIcoBut7 = Label(dashbrd, image = icoBut7, bg = cfg_clr)
        return
    genBut7(200, 60)
    def mouEnBut7(e):
        lbl7['font'] = ('Chirp', 14, 'bold')
    def mouLeBut7(e):
        lbl7['font'] = ('Chirp', 14)

    def genBut8(x, y):
        global icoBut8
        rawIcoBut8 = Image.open('./img/but.png')
        rawIcoBut8 = rawIcoBut8.resize((x, y), Image.Resampling.LANCZOS)
        icoBut8 = ImageTk.PhotoImage(rawIcoBut8)
        global lblIcoBut8
        lblIcoBut8 = Label(dashbrd, image = icoBut8, bg = cfg_clr)
        return
    genBut8(200, 60)
    def mouEnBut8(e):
        lbl8['font'] = ('Chirp', 14, 'bold')
    def mouLeBut8(e):
        lbl8['font'] = ('Chirp', 14)

    brand = Label(
        dashbrd,
        text = 'GRAND HOTEL',
        font = ('Montserrat', 18),
        foreground = 'black',
        bg = 'white'
    )
    brand.place(x=60, y=10)

    def genLogo(x, y):
        global logo
        rawLogo = Image.open('./img/logo.jpg')
        rawLogo = rawLogo.resize((x, y), Image.Resampling.LANCZOS)
        logo = ImageTk.PhotoImage(rawLogo)
        global lblLogo
        lblLogo = Label(dashbrd, image = logo, bg = cfg_clr)
        return
    genLogo(120, 120)
    lblLogo.place(x = 90, y = 60)

    lbl1 = Label(cursor = 'hand2', text = 'Phiếu thuê',bg = cfg_clr,font = ('Chirp', 14))
    lbl2 = Label(cursor = 'hand2', text = 'Hóa đơn',bg = cfg_clr,font = ('Chirp', 14))
    lbl3 = Label(cursor = 'hand2', text = 'Phòng',bg = cfg_clr,font = ('Chirp', 14))
    lbl4 = Label(cursor = 'hand2', text = 'Nhân viên',bg = cfg_clr,font = ('Chirp', 14))
    lbl5 = Label(cursor = 'hand2', text = 'Khách hàng' ,bg = cfg_clr,font = ('Chirp', 14))
    lbl6 = Label(cursor = 'hand2', text = 'Dịch vụ', bg = cfg_clr, font = ('Chirp', 14))
    lbl7 = Label(cursor = 'hand2', text = 'Kho hàng',bg = cfg_clr,font = ('Chirp', 14))
    lbl8 = Label(cursor = 'hand2', text = 'Thống kê',bg = cfg_clr,font = ('Chirp', 14))
    lbl9 = Label(cursor = 'hand2', text = 'Phiếu thuê',bg = cfg_clr,font = ('Chirp', 14, 'bold'))
    lbl10 = Label(cursor = 'hand2', text = 'Hóa đơn',bg = cfg_clr,font = ('Chirp', 14, 'bold'))
    lbl11 = Label(cursor = 'hand2', text = 'Phòng',bg = cfg_clr,font = ('Chirp', 14, 'bold'))
    lbl12 = Label(cursor = 'hand2', text = 'Nhân viên',bg = cfg_clr,font = ('Chirp', 14, 'bold'))
    lbl13 = Label(cursor = 'hand2', text = 'Khách hàng' ,bg = cfg_clr,font = ('Chirp', 14, 'bold'))
    lbl14 = Label(cursor = 'hand2', text = 'Dịch vụ',bg = cfg_clr,font = ('Chirp', 14, 'bold'))
    lbl15 = Label(cursor = 'hand2', text = 'Kho hàng',bg = cfg_clr,font = ('Chirp', 14, 'bold'))
    lbl16 = Label(cursor = 'hand2', text = 'Thống kê',bg = cfg_clr,font = ('Chirp', 14, 'bold'))


    LblPlacing = [30]
    for i in range(1,8):
        LblPlacing.append(LblPlacing[i-1]+50)

    def lblIcoButHovering():
        lblIcoBut1.bind("<Enter>", mouEnBut1)
        lblIcoBut1.bind("<Leave>", mouLeBut1)
        lblIcoBut2.bind("<Enter>", mouEnBut2)
        lblIcoBut2.bind("<Leave>", mouLeBut2)
        lblIcoBut3.bind("<Enter>", mouEnBut3)
        lblIcoBut3.bind("<Leave>", mouLeBut3)
        lblIcoBut4.bind("<Enter>", mouEnBut4)
        lblIcoBut4.bind("<Leave>", mouLeBut4)
        lblIcoBut5.bind("<Enter>", mouEnBut5)
        lblIcoBut5.bind("<Leave>", mouLeBut5)
        lblIcoBut6.bind("<Enter>", mouEnBut6)
        lblIcoBut6.bind("<Leave>", mouLeBut6)
        lblIcoBut7.bind("<Enter>", mouEnBut7)
        lblIcoBut7.bind("<Leave>", mouLeBut7)
        lblIcoBut8.bind("<Enter>", mouEnBut8)
        lblIcoBut8.bind("<Leave>", mouLeBut8)
    lblIcoButHovering()

    def toggleLbl1(e):
        lbl9.place(x=125, y=a[0]+10)
        lbl10.place_forget()
        lbl11.place_forget()
        lbl12.place_forget()
        lbl13.place_forget()
        lbl14.place_forget()
        lbl15.place_forget()
        lbl16.place_forget()
    def toggleLbl2(e):
        lbl10.place(x=125, y=a[1]+10)
        lbl9.place_forget()
        lbl11.place_forget()
        lbl12.place_forget()
        lbl13.place_forget()
        lbl14.place_forget()
        lbl15.place_forget()
        lbl16.place_forget()
    def toggleLbl3(e):
        lbl11.place(x=125, y=a[2]+10)
        lbl9.place_forget()
        lbl10.place_forget()
        lbl12.place_forget()
        lbl13.place_forget()
        lbl14.place_forget()
        lbl15.place_forget()
        lbl16.place_forget()
    def toggleLbl4(e):
        lbl12.place(x=125, y=a[3]+10)
        lbl9.place_forget()
        lbl10.place_forget()
        lbl11.place_forget()
        lbl13.place_forget()
        lbl14.place_forget()
        lbl15.place_forget()
        lbl16.place_forget()
    def toggleLbl5(e):
        lbl13.place(x=125, y=a[4]+10)
        lbl9.place_forget()
        lbl10.place_forget()
        lbl11.place_forget()
        lbl12.place_forget()
        lbl14.place_forget()
        lbl15.place_forget()
        lbl16.place_forget()
    def toggleLbl6(e):
        lbl14.place(x=125, y=a[5]+10)
        lbl9.place_forget()
        lbl10.place_forget()
        lbl11.place_forget()
        lbl12.place_forget()
        lbl13.place_forget()
        lbl15.place_forget()
        lbl16.place_forget()
    def toggleLbl7(e):
        lbl15.place(x=125, y=a[6]+10)
        lbl9.place_forget()
        lbl10.place_forget()
        lbl11.place_forget()
        lbl12.place_forget()
        lbl13.place_forget()
        lbl14.place_forget()
        lbl16.place_forget()
    def toggleLbl8(e):
        lbl16.place(x=125, y=a[7]+10)
        lbl9.place_forget()
        lbl10.place_forget()
        lbl11.place_forget()
        lbl12.place_forget()
        lbl13.place_forget()
        lbl14.place_forget()
        lbl15.place_forget()

    lftNavDiv.place(x=0, y=0) 
        
    def genIco1(x, y):
        global icoimg1
        rawIcoImg1 = Image.open('./img/formIcon.png')
        rawIcoImg1 = rawIcoImg1.resize((x, y), Image.Resampling.LANCZOS)
        icoimg1 = ImageTk.PhotoImage(rawIcoImg1)
        global lblIcoImg1
        lblIcoImg1 = Label(dashbrd, image = icoimg1, bg = cfg_clr, cursor = 'hand2')
        return
    def genIco2(x, y):
        global icoimg2
        rawIcoImg2 = Image.open('./img/billIcon.png')
        rawIcoImg2 = rawIcoImg2.resize((x, y), Image.Resampling.LANCZOS)
        icoimg2 = ImageTk.PhotoImage(rawIcoImg2)
        global lblIcoImg2
        lblIcoImg2 = Label(dashbrd, image = icoimg2, bg = cfg_clr, cursor = 'hand2')
        return
    def genIco3(x, y):
        global icoimg3
        rawIcoImg3 = Image.open('./img/room.png')
        rawIcoImg3 = rawIcoImg3.resize((x, y), Image.Resampling.LANCZOS)
        icoimg3 = ImageTk.PhotoImage(rawIcoImg3)
        global lblIcoImg3
        lblIcoImg3 = Label(dashbrd, image = icoimg3, bg = cfg_clr, cursor = 'hand2')
        return
    def genIco4(x, y):
        global icoimg4
        rawIcoImg4 = Image.open('./img/staff.png')
        rawIcoImg4 = rawIcoImg4.resize((x, y), Image.Resampling.LANCZOS)
        icoimg4 = ImageTk.PhotoImage(rawIcoImg4)
        global lblIcoImg4
        lblIcoImg4 = Label(dashbrd, image = icoimg4, bg = cfg_clr, cursor = 'hand2')
        return
    def genIco5(x, y):
        global icoimg5
        rawIcoImg5 = Image.open('./img/cust.png')
        rawIcoImg5 = rawIcoImg5.resize((x, y), Image.Resampling.LANCZOS)
        icoimg5 = ImageTk.PhotoImage(rawIcoImg5)
        global lblIcoImg5
        lblIcoImg5 = Label(dashbrd, image = icoimg5, bg = cfg_clr, cursor = 'hand2')
        return
    def genIco6(x, y):
        global icoimg6
        rawIcoImg6 = Image.open('./img/serv.png')
        rawIcoImg6 = rawIcoImg6.resize((x, y), Image.Resampling.LANCZOS)
        icoimg6 = ImageTk.PhotoImage(rawIcoImg6)
        global lblIcoImg6
        lblIcoImg6 = Label(dashbrd, image = icoimg6, bg = cfg_clr, cursor = 'hand2')
        return
    def genIco7(x, y):
        global icoimg7
        rawIcoImg7 = Image.open('./img/ware.png')
        rawIcoImg7 = rawIcoImg7.resize((x, y), Image.Resampling.LANCZOS)
        icoimg7 = ImageTk.PhotoImage(rawIcoImg7)
        global lblIcoImg7
        lblIcoImg7 = Label(dashbrd, image = icoimg7, bg = cfg_clr, cursor = 'hand2')
        return
    def genIco8(x, y):
        global icoimg8
        rawIcoImg8 = Image.open('./img/stat.png')
        rawIcoImg8 = rawIcoImg8.resize((x, y), Image.Resampling.LANCZOS)
        icoimg8 = ImageTk.PhotoImage(rawIcoImg8)
        global lblIcoImg8
        lblIcoImg8 = Label(dashbrd, image = icoimg8, bg = cfg_clr, cursor = 'hand2')
        return

    def lblIcoButPlacing():
        lblIcoBut1.place(x = 60, y = 195)
        lblIcoBut2.place(x = 60, y = 257)
        lblIcoBut3.place(x = 60, y = 317)
        lblIcoBut4.place(x = 60, y = 377)
        lblIcoBut5.place(x = 60, y = 437)
        lblIcoBut6.place(x = 60, y = 497)
        lblIcoBut7.place(x = 60, y = 557)
        lblIcoBut8.place(x = 60, y = 617)
    lblIcoButPlacing()

    def lblPlacing():
        lbl1.place(x=125, y=a[0]+10)
        lbl2.place(x=125, y=a[1]+10)
        lbl3.place(x=125, y=a[2]+10)
        lbl4.place(x=125, y=a[3]+10)
        lbl5.place(x=125, y=a[4]+10)
        lbl6.place(x=125, y=a[5]+10)
        lbl7.place(x=125, y=a[6]+10)
        lbl8.place(x=125, y=a[7]+10)
    lblPlacing()

    def genIco():
        genIco1(25, 25)
        genIco2(25, 30)
        genIco3(25, 30)
        genIco4(30, 30)
        genIco5(30, 30)
        genIco6(30, 30)
        genIco7(30, 30)
        genIco8(30, 30)
    genIco()

    def renderInpNhanVien(e):
        inpNv = Toplevel(dashbrd)
        inpNv.title('Điền thông tin nhân viên mới')
        inpNv.geometry('650x600')
        lbl1 = Label(inpNv, font = ('Chirp', 10), text = 'ID:')
        lbl2 = Label(inpNv, font = ('Chirp', 10), text = 'Họ')
        lbl3 = Label(inpNv, font = ('Chirp', 10), text = 'Tên')
        lbl4 = Label(inpNv, font = ('Chirp', 10), text = 'Giới tính')
        lbl5 = Label(inpNv, font = ('Chirp', 10), text = 'Năm sinh')
        lbl6 = Label(inpNv, font = ('Chirp', 10), text = 'Quê quán')
        lbl7 = Label(inpNv, font = ('Chirp', 10), text = 'Chức vụ')
        lbl8 = Label(inpNv, font = ('Chirp', 10), text = 'Bộ phận')
        lbl9 = Label(inpNv, font = ('Chirp', 10), text = 'Lương')
        lbl10 = Label(inpNv, font = ('Chirp', 10), text = 'Thưởng')
        heightSpacing = 50
        lbl1.place(x = 10, y = 10 + heightSpacing * 0)
        lbl2.place(x = 10, y = 10 + heightSpacing * 1)
        lbl3.place(x = 10, y = 10 + heightSpacing * 2)
        lbl4.place(x = 10, y = 10 + heightSpacing * 3)
        lbl5.place(x = 10, y = 10 + heightSpacing * 4)
        lbl6.place(x = 10, y = 10 + heightSpacing * 5)
        lbl7.place(x = 10, y = 10 + heightSpacing * 6)
        lbl8.place(x = 10, y = 10 + heightSpacing * 7)
        lbl9.place(x = 10, y = 10 + heightSpacing * 8)
        lbl10.place(x = 10, y = 10 + heightSpacing * 9)
         
        txt1Val = "M" + str(len(dataNv)+1).zfill(3)
        txt1 = Label(inpNv, font = ('Chirp', 10), height = 1, width = 40, text = txt1Val)
        txt2 = Text(inpNv, font = ('Chirp', 10), height = 1, width = 40)
        txt3 = Text(inpNv, font = ('Chirp', 10), height = 1, width = 40)

        genderEntry = ttk.Combobox(inpNv, values = ["Nam", "Nữ"], height = 1, width = 40)
        yearSample =  [*range(1950, 2020, 1)]
        txt5 = ttk.Combobox(inpNv, values = yearSample, font = ('Chirp', 10), height = 1, width = 40)
        txt6 = Text(inpNv, font = ('Chirp', 10), height = 1, width = 40)
        jobSample = []
        for row in dataDv:
            jobSample.append(row[6])
        txt7 = ttk.Combobox(inpNv, font = ('Chirp', 10), height = 1, width = 40, values = jobSample)
        txt8 = Text(inpNv, font = ('Chirp', 10), height = 1, width = 40)
        txt9 = Text(inpNv, font = ('Chirp', 10), height = 1, width = 40)
        txt10 = Text(inpNv, font = ('Chirp', 10), height = 1, width = 40)

        txt1.place(x = 100, y = 10 + heightSpacing * 0)
        txt2.place(x = 100, y = 10 + heightSpacing * 1)
        txt3.place(x = 100, y = 10 + heightSpacing * 2)
        genderEntry.place(x = 100, y = 10 + heightSpacing * 3)
        txt5.place(x = 100, y = 10 + heightSpacing * 4)
        txt6.place(x = 100, y = 10 + heightSpacing * 5)
        txt7.place(x = 100, y = 10 + heightSpacing * 6)
        txt8.place(x = 100, y = 10 + heightSpacing * 7)
        txt9.place(x = 100, y = 10 + heightSpacing * 8)
        txt10.place(x = 100, y = 10 + heightSpacing * 9)
        
        def addData():
        
            tmp = []
            tmp.append(txt1Val)
            tmp.append(txt2.get("1.0",'end-1c'))
            tmp.append(txt3.get("1.0",'end-1c'))
            tmp.append(genderEntry.get())
            tmp.append(int(txt5.get()))
            tmp.append(txt6.get("1.0",'end-1c'))
            tmp.append(txt7.get("1.0",'end-1c'))
            tmp.append(txt8.get("1.0",'end-1c'))
            try:
                tmp.append(int(txt9.get("1.0",'end-1c')))
            except ValueError:
                messagebox.showwarning(title = 'Cảnh báo', message = 'Vui lòng nhập tiền thưởng bằng số.')
            try:
                tmp.append(int(txt10.get("1.0",'end-1c')))
            except ValueError:
                messagebox.showwarning(title = 'Cảnh báo', message = 'Vui lòng nhập tiền thưởng bằng số.')
            global dataNv
            for row in dataNv:
                if (row[0] == tmp[0]):
                    messagebox.showwarning(title = 'Cảnh báo', message = 'ID này đã tồn tại.')
                    return
            if (len(tmp) != 10): return
            dataNv.append(tmp)
            messagebox.showwarning(title='Thông báo', message = 'Đã thêm thành công.')
            quitjob()
            genRight('nhanVien', True)
        def quitjob(): inpNv.destroy()
        but1 = Button(inpNv, text='Save', font = ('Chirp', 11), command =  addData)
        but2 = Button(inpNv, text='Cancel', font = ('Chirp', 11), command = quitjob)
        but1.place(x=120, y= 10 + heightSpacing * 10)
        but2.place(x=240, y= 10 + heightSpacing * 10)

    def renderInpPhieuThue(e):
        inpPTP = Toplevel(dashbrd)
        inpPTP.title('Phiếu thuê phòng')
        inpPTP.geometry('900x575')
        lbl1 = Label(inpPTP, font = ('Chirp', 10), text = 'ID phiếu')
        lbl2 = Label(inpPTP, font = ('Chirp', 10), text = 'ID khách hàng')
        lbl3 = Label(inpPTP, font = ('Chirp', 10), text = 'ID nhân viên')
        lbl8 = Label(inpPTP, font = ('Chirp', 10), text = 'ID Phòng')
        lbl4 = Label(inpPTP, font = ('Chirp', 10), text = 'ID dịch vụ')
        lbl5 = Label(inpPTP, font = ('Chirp', 10), text = 'Ngày đến')
        lbl6 = Label(inpPTP, font = ('Chirp', 10), text = 'Ngày đi')
        lbl7 = Label(inpPTP, font = ('Chirp', 10), text = 'Trả trước')
        heightSpacing = 50
        lbl1.place(x = 10, y = 10 + heightSpacing * 0)
        lbl2.place(x = 10, y = 10 + heightSpacing * 1)
        lbl3.place(x = 10, y = 10 + heightSpacing * 2)
        lbl4.place(x = 10, y = 10 + heightSpacing * 3)
        lbl5.place(x = 10, y = 10 + heightSpacing * 4)
        lbl6.place(x = 400, y = 10 + heightSpacing * 4)
        lbl7.place(x = 10, y = 10 + heightSpacing * 8)
        lbl8.place(x = 10, y = 10 + heightSpacing * 9)
        
        listkh = []
        for row in dataKh:
            listkh.append(row[0])
        listnv = []
        for row in dataNv:
            if (row[0][0] == 'M'):
                listnv.append(row[0])
        stt = "PTP" + str(len(dataPhieuThue)+1).zfill(4)
        txt1 = Label(inpPTP, font = ('Chirp', 10), height = 1, width = 40, text = stt)
        txt2 = ttk.Combobox(inpPTP, values = listkh, font = ('Chirp', 10), height = 1, width = 37)
        txt3 = ttk.Combobox(inpPTP, values = listnv, font = ('Chirp', 10), height = 1, width = 37)
        listdv = []
        for row in dataDv:
            listdv.append(row[0])
        txt5 = ttk.Combobox(inpPTP, values = listdv, font = ('Chirp', 10), height = 1, width = 37)

        txt6 = Calendar(inpPTP, selectmode = 'day', year = 2022, month = 5, day = 22, date_pattern="dd/mm/yyyy")
        txt7 = Calendar(inpPTP, selectmode = 'day', year = 2022, month = 5, day = 22, date_pattern="dd/mm/yyyy")
        i = IntVar()
        txt8 = Checkbutton(inpPTP, variable = i, text = 'Đã trả')
        listphg = []
        for row in dataPh:
            listphg.append(row[0])
        txt9 = ttk.Combobox(inpPTP, values = listphg, font = ('Chirp', 10), height = 1, width = 37)
        
        txt1.place(x = 100, y = 10 + heightSpacing * 0)
        txt2.place(x = 100, y = 10 + heightSpacing * 1)
        txt3.place(x = 100, y = 10 + heightSpacing * 2)
        txt5.place(x = 100, y = 10 + heightSpacing * 3)
        txt6.place(x = 100, y = 10 + heightSpacing * 4)
        txt7.place(x = 500, y = 10 + heightSpacing * 4)
        txt8.place(x = 100, y = 10 + heightSpacing * 8)
        txt9.place(x = 100, y = 10 + heightSpacing * 9)
        
        def addData():
            tmp = []
            tmp.append(stt)
            tmp.append(txt2.get())
            tmp.append(txt3.get())
            tmp.append(txt5.get())
            tmp1 = txt6.get_date()
            tmp2 = txt7.get_date()
            tmp3 = str(tmp1)
            tmp4 = str(tmp2)
            tmp.append(tmp3)
            tmp.append(tmp4)
            if (i.get() == 1):
                tmp.append(1)
            else:
                tmp.append(0)
            global dataPhieuThue
            tmp.append(0)
            tmp.append(0)
            tmp.append(0)
            for row in dataPh:
                if (row[0] == txt9.get() and row[2] == 'maintaining'): 
                    messagebox.showwarning(title="Cảnh báo", message = "Phòng đang bảo trì")
                    return
            
            tmp.append(txt9.get())
            dataPhieuThue.append(tmp)
            messagebox.showwarning(title='Thông báo', message = 'Đã thêm thành công.')
            inpPTP.destroy()
            genRight('phieuThue', False)
            return
        def quitjob():
            inpPTP.destroy()
            return
        but1 = Button(inpPTP, text='Save', font = ('Chirp', 11), command =  addData)
        but2 = Button(inpPTP, text='Cancel', font = ('Chirp', 11), command = quitjob)
        but1.place(x=120, y= 10 + heightSpacing * 10)
        but2.place(x=240, y= 10 + heightSpacing * 10)
        return

    def renderInpPhong(e):
        phg = Toplevel(dashbrd)
        phg.title('Thông tin phòng')
        phg.geometry('400x400')
        lbl1 = Label(phg, font = ('Chirp', 10), text = 'ID phòng')
        lbl2 = Label(phg, font = ('Chirp', 10), text = 'Giá phòng')
        lbl3 = Label(phg, font = ('Chirp', 10), text = 'Trang thái')
        lbl4 = Label(phg, font = ('Chirp', 10), text = 'Mô tả')
        lbl5 = Label(phg, font = ('Chirp', 10), text = 'Kiểu phòng')
        lbl6 = Label(phg, font = ('Chirp', 10), text = 'Trang bị')
        heightSpacing = 50
        lbl1.place(x = 10, y = 10 + heightSpacing * 0)
        lbl2.place(x = 10, y = 10 + heightSpacing * 1)
        lbl3.place(x = 10, y = 10 + heightSpacing * 2)
        lbl4.place(x = 10, y = 10 + heightSpacing * 3)
        lbl5.place(x = 10, y = 10 + heightSpacing * 4)
        lbl6.place(x = 10, y = 10 + heightSpacing * 5)
         
        txt1 = Text(phg, font = ('Chirp', 10), height = 1, width = 40)
        txt2 = Text(phg, font = ('Chirp', 10), height = 1, width = 40)
        txt3 = Text(phg, font = ('Chirp', 10), height = 1, width = 40)
        available = ttk.Combobox(phg, values = ['available', 'maintaining', 'occupied'])
        var2 = ttk.Combobox(phg, values = ['normal', 'vip', 'elite'])

        txt5 = Text(phg, font = ('Chirp', 10), height = 1, width = 40)
        txt6 = Text(phg, font = ('Chirp', 10), height = 1, width = 40)
        

        txt1.place(x = 100, y = 10 + heightSpacing * 0)
        txt2.place(x = 100, y = 10 + heightSpacing * 1)
        available.place(x = 100, y = 10 + heightSpacing * 2)
        var2.place(x = 100, y = 10 + heightSpacing * 4)
        txt5.place(x = 100, y = 10 + heightSpacing * 3)
        txt6.place(x = 100, y = 10 + heightSpacing * 5)
        
        def addData():
            tmp = []
            tmp.append(txt1.get("1.0",'end-1c'))
            tmp.append(int(txt2.get("1.0",'end-1c')))
            tmp.append(available.get())
            tmp.append(txt5.get("1.0",'end-1c'))
            tmp.append(var2.get())
            tmp.append(txt6.get("1.0",'end-1c'))
            global dataPh
            dataPh.append(tmp)
            genRight('phong', False)
            phg.destroy()
        def quitjob(): phg.destroy()
        but1 = Button(phg, text='Save', font = ('Chirp', 11), command =  addData)
        but2 = Button(phg, text='Cancel', font = ('Chirp', 11), command = quitjob)
        but1.place(x=120, y= 10 + heightSpacing * 6)
        but2.place(x=240, y= 10 + heightSpacing * 6)

        return

    def renderInpKh(e):
        inpKh = Toplevel(dashbrd)
        inpKh.title('Thông tin khách hàng')
        inpKh.geometry('400x380')
        lbl1 = Label(inpKh, font = ('Chirp', 10), text = 'ID')
        lbl2 = Label(inpKh, font = ('Chirp', 10), text = 'Họ')
        lbl3 = Label(inpKh, font = ('Chirp', 10), text = 'Tên')
        lbl4 = Label(inpKh, font = ('Chirp', 10), text = 'Giới tính')
        lbl5 = Label(inpKh, font = ('Chirp', 10), text = 'Năm sinh')
        lbl6 = Label(inpKh, font = ('Chirp', 10), text = 'Địa chỉ')
        
        heightSpacing = 50
        lbl1.place(x = 10, y = 10 + heightSpacing * 0)
        lbl2.place(x = 10, y = 10 + heightSpacing * 1)
        lbl3.place(x = 10, y = 10 + heightSpacing * 2)
        lbl4.place(x = 10, y = 10 + heightSpacing * 3)
        lbl5.place(x = 10, y = 10 + heightSpacing * 4)
        lbl6.place(x = 10, y = 10 + heightSpacing * 5)
        
         
        txt1 = Text(inpKh, font = ('Chirp', 10), height = 1, width = 40)
        txt2 = Text(inpKh, font = ('Chirp', 10), height = 1, width = 40)
        txt3 = Text(inpKh, font = ('Chirp', 10), height = 1, width = 40)

        var = StringVar()
        gender = ttk.Combobox(inpKh, values = ['Nam', 'Nữ'])
        txt5 = Text(inpKh, font = ('Chirp', 10), height = 1, width = 40)
        txt6 = Text(inpKh, font = ('Chirp', 10), height = 1, width = 40)
        

        txt1.place(x = 100, y = 10 + heightSpacing * 0)
        txt2.place(x = 100, y = 10 + heightSpacing * 1)
        txt3.place(x = 100, y = 10 + heightSpacing * 2)
        gender.place(x = 100, y = 10 + heightSpacing * 3)
        txt5.place(x = 100, y = 10 + heightSpacing * 4)
        txt6.place(x = 100, y = 10 + heightSpacing * 5)
        
        def addData():
            tmp = []
            tmp.append(txt1.get("1.0",'end-1c'))
            tmp.append(txt2.get("1.0",'end-1c'))
            tmp.append(txt3.get("1.0",'end-1c'))
            tmp.append(gender.get())
            tmp.append(int(txt5.get("1.0",'end-1c')))
            tmp.append(txt6.get("1.0",'end-1c'))
            for i in range(0, 4): tmp.append('')
            global dataKh
            dataKh.append(tmp)
            genRight('khachHang', False)
        def quitjob(): inpKh.destroy()

        but1 = Button(inpKh, text='Save', font = ('Chirp', 11), command =  addData)
        but2 = Button(inpKh, text='Cancel', font = ('Chirp', 11), command = quitjob)
        but1.place(x=120, y= 10 + heightSpacing * 6)
        but2.place(x=240, y= 10 + heightSpacing * 6)
        return

    def renderInpDv(e):
        inpDv = Toplevel(dashbrd)
        inpDv.title('Thông tin dịch vụ')
        inpDv.geometry('400x200')
        lbl1 = Label(inpDv, font = ('Chirp', 10), text = 'ID')
        lbl2 = Label(inpDv, font = ('Chirp', 10), text = 'Tên dịch vụ')
        lbl3 = Label(inpDv, font = ('Chirp', 10), text = 'Chi phí')

        heightSpacing = 50
        lbl1.place(x = 10, y = 10 + heightSpacing * 0)
        lbl2.place(x = 10, y = 10 + heightSpacing * 1)
        lbl3.place(x = 10, y = 10 + heightSpacing * 2)

        txt1 = Text(inpDv, font = ('Chirp', 10), height = 1, width = 40)
        txt2 = Text(inpDv, font = ('Chirp', 10), height = 1, width = 40)
        txt3 = Text(inpDv, font = ('Chirp', 10), height = 1, width = 40)

        txt1.place(x = 100, y = 10 + heightSpacing * 0)
        txt2.place(x = 100, y = 10 + heightSpacing * 1)
        txt3.place(x = 100, y = 10 + heightSpacing * 2)
        global dataDv
        txt1.insert('end', 'DV' + str(len(dataDv)+1).zfill(3))

        def addData():
            tmp = []
            global dataDv
            tmp.append(txt1.get("1.0",'end-1c'))
            tmp.append(txt2.get("1.0",'end-1c'))
            try:
                tmp.append(int(txt3.get("1.0",'end-1c')))
            except ValueError:
                messagebox.showwarning(title='Cảnh báo', message = 'Vui lòng nhập giá dịch vụ ở định dạng số.')
                return
            for row in dataDv:
                if (row[0] == txt1.get("1.0",'end-1c')):
                    messagebox.showwarning(title='Cảnh báo', message = 'ID dịch vụ đã tồn tại. Vui lòng dùng ID khác.')
                    return
            dataDv.append(tmp)
            genRight('dichVu', False)
            inpDv.destroy()
        def quitjob(): inpDv.destroy()
        but1 = Button(inpDv, text='Save', font = ('Chirp', 11), command =  addData)
        but2 = Button(inpDv, text='Cancel', font = ('Chirp', 11), command = quitjob)
        but1.place(x=120, y= 10 + heightSpacing * 3)
        but2.place(x=240, y= 10 + heightSpacing * 3)
        return

    def renderInpProduct(e):
        inpPr = Toplevel(dashbrd)
        inpPr.title('Thêm thông tin thiết bị và thực phẩm')
        inpPr.geometry('700x700')
        lbl1 = Label(inpPr, font = ('Chirp', 10), text = 'ID phiếu')
        lbl2 = Label(inpPr, font = ('Chirp', 10), text = 'ID hàng hóa')
        lbl3 = Label(inpPr, font = ('Chirp', 10), text = 'Tên hàng hóa')
        lbl4 = Label(inpPr, font = ('Chirp', 10), text = 'Số lượng')
        lbl5 = Label(inpPr, font = ('Chirp', 10), text = 'ID nhà cung cấp')
        lbl6 = Label(inpPr, font = ('Chirp', 10), text = 'Tên nhà cung cấp')
        lbl7 = Label(inpPr, font = ('Chirp', 10), text = 'Nhà vận chuyển')
        lbl8 = Label(inpPr, font = ('Chirp', 10), text = 'Phí vận chuyển')
        lbl9 = Label(inpPr, font = ('Chirp', 10), text = 'Ngày nhập')

        heightSpacing = 50
        lbl1.place(x = 10, y = 10 + heightSpacing * 0)
        lbl2.place(x = 10, y = 10 + heightSpacing * 1)
        lbl3.place(x = 10, y = 10 + heightSpacing * 2)
        lbl4.place(x = 10, y = 10 + heightSpacing * 3)
        lbl5.place(x = 10, y = 10 + heightSpacing * 4)
        lbl6.place(x = 10, y = 10 + heightSpacing * 5)
        lbl7.place(x = 10, y = 10 + heightSpacing * 6)
        lbl8.place(x = 10, y = 10 + heightSpacing * 7)
        lbl9.place(x = 10, y = 10 + heightSpacing * 8)

        idSuggestion = "P" + str(len(phieuNhapTbAndFood)+1).zfill(3)
        txt1 = Label(inpPr, text = idSuggestion, font = ('Chirp', 10), height = 1, width = 40)
        txt2 = Text(inpPr, font = ('Chirp', 10), height = 1, width = 40)
        txt3 = Text(inpPr, font = ('Chirp', 10), height = 1, width = 40)
        amount = [*range (1, 300)]
        txt4 = ttk.Combobox(inpPr, values = amount, font = ('Chirp', 10), height = 1, width = 37)
        
        supplier = []
        for item in phieuNhapTbAndFood:
            supplier.append(item[5])
        txt5 = ttk.Combobox(inpPr, values = supplier, font = ('Chirp', 10), height = 1, width = 40)
        txt6 = Text(inpPr, font = ('Chirp', 10), height = 1, width = 40)
        txt7 = Text(inpPr, font = ('Chirp', 10), height = 1, width = 40)
        txt8 = Text(inpPr, font = ('Chirp', 10), height = 1, width = 40)
        today = datetime.today().strftime('%d-%m-%Y')
        txt9 = Label(inpPr, text = today, font = ('Chirp', 10), height = 1, width = 40)

        txt1.place(x = 125, y = 10 + heightSpacing * 0)
        txt2.place(x = 125, y = 10 + heightSpacing * 1)
        txt3.place(x = 125, y = 10 + heightSpacing * 2)
        txt4.place(x = 125, y = 10 + heightSpacing * 3)
        txt5.place(x = 125, y = 10 + heightSpacing * 4)
        txt6.place(x = 125, y = 10 + heightSpacing * 5)
        txt7.place(x = 125, y = 10 + heightSpacing * 6)
        txt8.place(x = 125, y = 10 + heightSpacing * 7)
        txt9.place(x = 125, y = 10 + heightSpacing * 8)
        def addData():
            tmp = []
            tmp.append(idSuggestion)
            tmp.append(txt2.get("1.0",'end-1c'))
            tmp.append(txt3.get("1.0",'end-1c'))
            tmp.append(txt4.get())
            tmp.append(txt5.get("1.0",'end-1c'))
            tmp.append(txt6.get("1.0",'end-1c'))
            tmp.append(txt7.get("1.0",'end-1c'))
            tmp.append(txt8.get("1.0",'end-1c'))
            tmp.append(today)
            global phieuNhapTbAndFood
            phieuNhapTbAndFood.append(tmp)
            genRight('pNhapTbiAndFood', False)
            return
        def quitjob(): inpPr.destroy()
        but1 = Button(inpPr, text='Save', font = ('Chirp', 11), command =  addData)
        but2 = Button(inpPr, text='Cancel', font = ('Chirp', 11), command = quitjob)
        but1.place(x=120, y= 10 + heightSpacing * 9)
        but2.place(x=240, y= 10 + heightSpacing * 9)
        return

    refreshBut = Button(dashbrd, text = 'Refresh', bg = 'white')
    refreshBut.place(x = 960, y = 105)
    deleteBut = Button(dashbrd, text = 'Delete', bg = 'white')
    deleteBut.place(x = 1020, y = 105)

    def deleteRow(e):
        if (chungTaDangODau == 'phieuThanhToan'):
            messagebox.showwarning(title='Cảnh báo', message='Chức năng không hỗ trợ.')
            return
        delFr = Toplevel(dashbrd)
        delFr.title('Xóa item')
        delFr.geometry('300x180')
        text = Label(delFr, text = 'Nhập stt hàng cần xóa: ', padx = 5, pady = 5)
        text.place(x = 10, y = 10)
        inp = Text(delFr, width = 30, padx = 5, pady = 5, height = 1)
        inp.place(x = 10, y = 50)

        def deleteAndExit():
            k = -1
            try:
                k = int(inp.get("1.0",'end-1c'))
            except:
                messagebox.showwarning(title='Cảnh báo', message='Vui lòng nhập đúng định dạng.')
            if (chungTaDangODau == 'phieuThue'):
                global dataPhieuThue
                del dataPhieuThue[k-1]
                genRight('phieuThue', False)
            if (chungTaDangODau == 'nhanVien'):
                global dataNv
                del dataNv[k-1]
                genRight('nhanVien', False)
            if (chungTaDangODau == 'phong'):
                global dataPh
                del dataPh[k-1]
                genRight('phong', False)
            if (chungTaDangODau == 'khachHang'):
                global dataKh
                del dataKh[k-1]
                genRight('khachHang', False)
            if (chungTaDangODau == 'dichVu'):
                global dataDv
                del dataDv[k-1]
                genRight('dichVu', False)
            if (chungTaDangODau == 'pNhapTbiAndFood'):
                global phieuNhapTbAndFood
                del phieuNhapTbAndFood[k-1]
                genRight('pNhapTbiAndFood', False)
            inp.destroy()
        save = Button(delFr, text = 'Delete', command = deleteAndExit)
        save.place(x = 120, y = 100)
        return
    deleteBut.bind("<Button-1>", deleteRow)
    
    
    def notavailable():
        messagebox.showwarning(title='Thông báo', message='Hóa đơn tự động sinh từ phiếu thuê phòng.')
        return

    def temp1(e):
        global tabSel
        tabSel = [False] * 7
        tabSel[0] = True, toggleLbl1
        genRight('phieuThue', False), genBotBut('phieuThue')
        refreshBut.bind("<Button-1>", temp1)
        addBut1.bind("<Button-1>", renderInpPhieuThue)
    def temp2(e): 
        global tabSel
        tabSel = [False] * 7
        tabSel[1] = True, toggleLbl2
        genRight('phieuThanhToan', False), genBotBut('phieuThanhToan')
        refreshBut.bind("<Button-1>", temp2)
        addBut1.bind("<Button-1>", notavailable)
    def temp3(e): 
        global tabSel
        tabSel = [False] * 7
        tabSel[2] = True
        toggleLbl3
        genRight('phong', False), genBotBut('phong')
        refreshBut.bind("<Button-1>", temp3)
        addBut1.bind("<Button-1>", renderInpPhong)
    def temp4(e): 
        global tabSel
        tabSel = [False] * 7
        tabSel[3] = True
        toggleLbl4
        genRight('nhanVien', False), genBotBut('nhanVien')
        refreshBut.bind("<Button-1>", temp4)
        addBut1.bind("<Button-1>", renderInpNhanVien)
    def temp5(e): 
        global tabSel
        tabSel = [False] * 7
        tabSel[4] = True
        toggleLbl5
        genRight('khachHang', False), genBotBut('khachHang')
        refreshBut.bind("<Button-1>", temp5)
        addBut1.bind("<Button-1>", renderInpKh)
    def temp6(e): 
        global tabSel
        tabSel = [False] * 7
        tabSel[5] = True
        toggleLbl6
        genRight('dichVu', False), genBotBut('dichVu')
        refreshBut.bind("<Button-1>", temp6)
        addBut1.bind("<Button-1>", renderInpDv)

    def temp7(e): 
        global tabSel
        tabSel = [False] * 7
        tabSel[6] = True
        toggleLbl7
        genRight('pNhapTbiAndFood', False), genBotBut('pNhapTbiAndFood')
        refreshBut.bind("<Button-1>", temp7)
        addBut1.bind("<Button-1>", renderInpProduct)

    def thongke():
        fig = matplotlib.figure.Figure(figsize=(5,5))
        ax = fig.add_subplot(111)
        percent = []
        label = []
        for row in phieuNhapTbAndFood:
            percent.append(row[3])
            label.append(row[2])
        ax.pie(percent) 
        ax.legend(label)
        circle=matplotlib.patches.Circle((0,0), 0.7, color='white')
        ax.add_artist(circle)

        thongke = Toplevel(dashbrd)
        thongke.title('Thống kê cơ cấu số lượng sản phầm tồn kho')
        canvas = FigureCanvasTkAgg(fig, master = thongke)
        canvas.get_tk_widget().pack()
        canvas.draw()
        return
    def temp8(e): 
        global tabSel
        tabSel = [False] * 7
        toggleLbl8
        thongke()
        # genRight('phieuThanhToan', False), genBotBut('phieuThanhToan')
        # refreshBut.bind("<Button-1>", temp8)



    addBut1 = Button(
        dashbrd,
        image = icoAdd,
        font = ('Chirp', 14),
        bg = 'white',
        border = '1px solid black',
        # text = "Add"
    )
    addBut1.bind("<Button-1>", renderInpPhieuThue)
    addBut1.place(x=340, y=110)

    def onclickEffectAndGuide():
        

        lbl1.bind("<Button-1>", temp1)
        lbl2.bind("<Button-1>", temp2)
        lbl3.bind("<Button-1>", temp3)
        lbl4.bind("<Button-1>", temp4)
        lbl5.bind("<Button-1>", temp5)
        lbl6.bind("<Button-1>", temp6)
        lbl7.bind("<Button-1>", temp7)
        lbl8.bind("<Button-1>", temp8)
        lblIcoBut1.bind("<Button-1>", temp1)
        lblIcoBut2.bind("<Button-1>", temp2)
        lblIcoBut3.bind("<Button-1>", temp3)
        lblIcoBut4.bind("<Button-1>", temp4)
        lblIcoBut5.bind("<Button-1>", temp5)
        lblIcoBut6.bind("<Button-1>", temp6)
        lblIcoBut7.bind("<Button-1>", temp7)
        lblIcoBut8.bind("<Button-1>", temp8)

        lblIcoImg1.bind("<Button-1>", temp1)
        lblIcoImg2.bind("<Button-1>", temp2)
        lblIcoImg3.bind("<Button-1>", temp3)
        lblIcoImg4.bind("<Button-1>", temp4)
        lblIcoImg5.bind("<Button-1>", temp5)
        lblIcoImg6.bind("<Button-1>", temp6)
        lblIcoImg7.bind("<Button-1>", temp7)
        lblIcoImg8.bind("<Button-1>", temp8)
    onclickEffectAndGuide()

    def hovering():
        lblIcoImg1.bind("<Enter>", mouEnBut1)
        lblIcoImg1.bind("<Leave>", mouLeBut1)
        lbl1.bind("<Enter>", mouEnBut1)
        lbl1.bind("<Leave>", mouLeBut1)
        lblIcoImg2.bind("<Enter>", mouEnBut2)
        lblIcoImg2.bind("<Leave>", mouLeBut2)
        lbl2.bind("<Enter>", mouEnBut2)
        lbl2.bind("<Leave>", mouLeBut2)
        lblIcoImg3.bind("<Enter>", mouEnBut3)
        lblIcoImg3.bind("<Leave>", mouLeBut3)
        lbl3.bind("<Enter>", mouEnBut3)
        lbl3.bind("<Leave>", mouLeBut3)
        lblIcoImg4.bind("<Enter>", mouEnBut4)
        lblIcoImg4.bind("<Leave>", mouLeBut4)
        lbl4.bind("<Enter>", mouEnBut4)
        lbl4.bind("<Leave>", mouLeBut4)
        lblIcoImg5.bind("<Enter>", mouEnBut5)
        lblIcoImg5.bind("<Leave>", mouLeBut5)
        lbl5.bind("<Enter>", mouEnBut5)
        lbl5.bind("<Leave>", mouLeBut5)
        lblIcoImg6.bind("<Enter>", mouEnBut6)
        lblIcoImg6.bind("<Leave>", mouLeBut6)
        lbl6.bind("<Enter>", mouEnBut6)
        lbl6.bind("<Leave>", mouLeBut6)
        lblIcoImg7.bind("<Enter>", mouEnBut7)
        lblIcoImg7.bind("<Leave>", mouLeBut7)
        lbl7.bind("<Enter>", mouEnBut7)
        lbl7.bind("<Leave>", mouLeBut7)
        lblIcoImg8.bind("<Enter>", mouEnBut8)
        lblIcoImg8.bind("<Leave>", mouLeBut8)
        lbl8.bind("<Enter>", mouEnBut8)
        lbl8.bind("<Leave>", mouLeBut8)
    hovering()

    def placing():
        lblIcoImg1.place(x = 90, y = a[0]+10)
        lblIcoImg2.place(x = 90, y = a[1]+10)
        lblIcoImg3.place(x = 90, y = a[2]+10)
        lblIcoImg4.place(x = 87.5, y = a[3]+7.5)
        lblIcoImg5.place(x = 87.5, y = a[4]+7.5)
        lblIcoImg6.place(x = 87.5, y = a[5]+7.5)
        lblIcoImg7.place(x = 87.5, y = a[6]+7.5)
        lblIcoImg8.place(x = 87.5, y = a[7]+7.5)
    placing()


def updateDb():
    print('update db')
    return

def startAutoSave():
    global thrd1
    thrd1 = threading.Thread(target = autosave, args = ())
    thrd1.start()
    return

def mouOnBut(k):
    if (k==1):
        print('render phieu')
    return

def autosave():
    print('auto saving...')
    time.sleep(2)
    if (winClosed):
        try:
            thrd1.join()
        except (RuntimeError):
            pass
        return
    else:
        autosave()

def startKeyListener():
    def on_press(key):
        if (winClosed):
            return False
        if key == keyboard.Key.enter:
            updateDb()
        try:
            k = key.char 
        except:
            k = key.name 
        
    global listener
    listener = keyboard.Listener(on_press=on_press)
    listener.start()  

def genDashUI():
    dashbrd.title('Dashboard')
    dashbrd.iconbitmap(appIco)
    dashbrd.configure(bg='white')

    tmp = str(appWidth) + 'x' + str(appHeight) + '+' + str(60) + '+' + str(0)
    dashbrd.geometry(tmp)
    dashbrd.resizable(False, False)


def genTopBanner():

    # global topimg
    # rawtopimg = Image.open('./img/hotel.jpg')
    # rawtopimg = rawtopimg.resize((600, 200), Image.Resampling.LANCZOS)
    # topimg = ImageTk.PhotoImage(rawtopimg)
    # lblTopImg = Label(dashbrd, image = topimg)
    # lblTopImg.place(x = 0, y = 0, relx = .25, rely = .25)

    global hdr_img
    raw_hdr_img = Image.open('./img/wall.png')
    raw_hdr_img = raw_hdr_img.resize((900, 787), Image.Resampling.LANCZOS)
    hdr_img = ImageTk.PhotoImage(raw_hdr_img)
    lbl_hdr_img = Label(dashbrd, image = hdr_img)

    lbl_hdr_img.place(x = 301, y = 0)

def genRight(s, includeSearch):
    
    global dvHeader
    dvHeader = ['ID', 'Tên dịch vụ', 'Chi phí']
    
    global phongHeader
    phongHeader = ['ID', 'Giá phòng', 'Trạng thái', 'Mô tả', 'Kiểu phòng', 'Trang bị']
    
    global ptpHeader
    ptpHeader = ["ID Phiếu", "ID Khách hàng", "ID Nhân viên", "Danh sách dịch vụ", "Ngày ở", "Ngày đến", "Trả trước", "Tiền thuê", "Phí dịch vụ", "Tổng tiền", "ID Phòng"]
    
    global pttHeader
    pttHeader = ['ID Phiếu', 'ID dịch vụ', 'ID nhân viên', 'Tên nhân viên', 'Số ngày ở', 'Tổng tiền', 'VAT', 'Tiền phải trả', 'Ngày in' ]
    
    # generate bill from pre-paid bill 
    dataPhieuTT = []
    for row in dataPhieuThue:
        index = dataPhieuThue.index(row)
        tmp = []
        tmp.append('PTT' + row[0][3:])
        tmp.append(row[3])
        tmp.append(row[2])
        for item in dataNv:
            if (item[0] == row[2]):
                tmp.append(item[1]+ ' ' +item[2])
                break
        
        d0 = date(int(row[4][6:10]), int(row[4][3:5]), int(row[4][0:2]))
        d1 = date(int(row[5][6:10]), int(row[5][3:5]), int(row[5][0:2]))
        delta = d1 - d0
        tmp.append(delta.days)

        rentFee = 0
        if (row[10] != ''):
            for row1 in dataPh: 
                if (row1[0] == row[10]):
                    rentFee = int(row1[1])
                    break
        serviceFee = 0 
        if (row[3] != ''):
            for row2 in dataDv:
                if (row2[0] == row[3]):
                    serviceFee = int(row2[2])
        tmp.append(str(rentFee+serviceFee))
        tmp.append('.15')
        tmp.append(
            str(
                round((rentFee+serviceFee)*1.15, 1)
            )
        )
        tmp.append(str(date.today().isoformat()))

        dataPhieuThue[index][7] = rentFee
        dataPhieuThue[index][8] = serviceFee
        dataPhieuThue[index][9] = rentFee + serviceFee

        dataPhieuTT.append(tmp)
    
    dct = {
        'nhanVien': dataNv,
        'khachHang': dataKh,
        'phieuThue': dataPhieuThue,
        'phieuThanhToan': dataPhieuTT,
        'pNhapTbiAndFood': phieuNhapTbAndFood,
        'phong': dataPh,
        'dichVu': dataDv
    }

    with open('./json/data.json', 'w', encoding='utf-8') as fi:
        json.dump(dct, fi, ensure_ascii=False, indent=4)

    with open('./json/data.json', 'r', encoding='utf-8') as fo:
        dataRead = json.loads(fo.read())
    
    # print(json.dumps(dataRead['khachHang'], ensure_ascii=False, indent = 4))
    dataTbl = json.dumps(dataRead[s], ensure_ascii=False)
    global dataTblList
    if (not includeSearch):
        dataTblList = json.loads(dataTbl)

    global nvHeader
    nvHeader = ["ID", "Họ", "Tên", "Giới tính", "Năm sinh", "Quê quán", "Chức vụ", "Bộ phận", "Lương", "Thưởng"]
    global khHeader
    khHeader = ['ID', 'Họ', 'Tên', 'Giới tính', 'Năm sinh', 'Địa chỉ', 'Số CMND/CDDD', 'Quốc tịch', 'Số điện thoại', 'Email']
    global pnHeader
    pnHeader = ['ID phiếu', 'ID Hàng hóa', 'Tên hàng hóa', 'Số lượng', 'ID nhà cung cấp', 'Tên nhà cung cấp', 'ID nhân viên', 'Phí vận chuyển', 'Ngày nhập']
    
    
    global toRenderHeader
    toRenderHeader = []
    if (s == 'nhanVien'):
        toRenderHeader = nvHeader
    elif (s == 'phieuThue'):
        toRenderHeader = ptpHeader
    elif (s == 'khachHang'):
        toRenderHeader = khHeader
    elif (s == 'phieuThanhToan'):
        toRenderHeader = pttHeader
        tmpnvTbl = json.loads(json.dumps(dataRead['nhanVien'], ensure_ascii=False))
        tmpDataTbl = dataTblList
        for r1 in tmpDataTbl:
            isFound = False
            for r2 in tmpnvTbl:
                if (r2[0] == r1[2]):
                    r1[3] = r2[1] + r2[2]
                    isFound = True
            if (not isFound):
                r1[3] = '<Not found>'
    elif (s == 'pNhapTbiAndFood'):
        toRenderHeader = pnHeader
    elif (s == 'dichVu'):
        toRenderHeader = dvHeader
    elif (s == 'phong'):
        toRenderHeader = phongHeader
    global rSh
    rSh = Sheet(dashbrd, 
        show_table = True,
        width = tblWidth,
        height = tblHeight,
        show_header = True,
        row_height = 70,
        data = dataTblList,
        headers = list(toRenderHeader)
    )
    # rSh.set_cell_data(0, 0, value = 999, set_copy = True, redraw = False)
    rSh.set_all_cell_sizes_to_text(redraw = True)
    rSh.enable_bindings('all')
    rSh.place(x=340, y=180)

def genBotBut(strTable):
    global chungTaDangODau
    chungTaDangODau = strTable
    with open('./json/data.json', 'r', encoding='utf-8') as fo:
        dataRead = json.loads(fo.read())
    dataTbl4GenRight = json.dumps(dataRead[strTable], ensure_ascii=False)
    dataTblList4GenRight = json.loads(dataTbl4GenRight)
    # print(dataTblList4GenRight[0][0])

    def exportXlsx():
        cur = con.cursor()
        cur.execute("SELECT * FROM {}".format(strTable))
        con.commit()
        resultToPrint = cur.fetchall()
        if (strTable == 'phieuThue'):
            tableDataFrame = pandas.DataFrame(resultToPrint, columns = ptpHeader)
        elif (strTable == 'phieuThanhToan'):
            tableDataFrame = pandas.DataFrame(resultToPrint, columns = pttHeader)
        elif (strTable == 'nhanVien'):
            tableDataFrame = pandas.DataFrame(resultToPrint, columns = nvHeader)
        elif (strTable == 'dichVu'):
            tableDataFrame = pandas.DataFrame(resultToPrint, columns = dvHeader)
        elif (strTable == 'khachHang'):
            tableDataFrame = pandas.DataFrame(resultToPrint, columns = khHeader)
        elif (strTable == 'phong'):
            tableDataFrame = pandas.DataFrame(resultToPrint, columns = phongHeader)
        elif (strTable == 'pNhapTbiAndFood'):
            tableDataFrame = pandas.DataFrame(resultToPrint, columns = pnHeader)
        
        tableDataFrame.to_excel('qlks.xlsx', sheet_name='sheet1', index=False)
        wb = openpyxl.load_workbook(filename = "qlks.xlsx")
        worksheet = wb.active
        for column in ascii_uppercase:
            worksheet.column_dimensions[column].width = 25
        wb.save('qlks.xlsx')
        messagebox.showwarning(title = "Thông báo", message = "Đã export thành công.")
        return

    saveBotBut = Button(
        dashbrd,
        font = ('Chirp', 14),
        text = 'Save ',
        bg = 'white',
        border = '1px solid black',
        height = 1,
        width = 7
    )
    exportXlsxBotBut = Button(
        dashbrd,
        font = ('Chirp', 14),
        text = 'Export Xlsx',
        bg = 'white',
        border = '1px solid black',
        height = 1,
        width = 20,
        command = exportXlsx
    )
    def printSelectedRow():
        data2Print = []
        if (strTable != 'phieuThanhToan'):
            messagebox.showwarning(title='Cảnh báo', message='Tính năng không khả dụng. Vui lòng chọn mục Hóa đơn để in.')
            return
        a = rSh.get_currently_selected(get_coords = False, return_nones_if_not = False)
        try:
            if (a[0] == 'row'):
                data2Print = dataTblList[a[1]]
                index = a[1]
            else:
                data2Print = dataTblList[a[0]]
                index = a[0]
        except (IndexError):
            messagebox.showwarning(title = 'Cảnh báo', message = 'Vui lòng chọn hàng cần xuất.')

        pdf = FPDF()
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.add_font("NotoSans", style="", fname="./fonts/NotoSans-Regular.ttf", uni=True)
        pdf.add_font("NotoSans", style="B", fname="./fonts/NotoSans-Bold.ttf", uni=True)
        pdf.add_font("NotoSans", style="I", fname="./fonts/NotoSans-Italic.ttf", uni=True)
        pdf.add_font("NotoSans", style="BI", fname="./fonts/NotoSans-BoldItalic.ttf", uni=True)
        
        pdf.set_font("NotoSans", style="B", size=12)
        pdf.cell(0, 15, 'KHÁCH SẠN GRAND HOTEL', 0, 1, 'C')
        pdf.set_font("NotoSans", style="", size=12)
        #image
        pdf.cell(0, 10, '84/176 Phan Văn Trị, P.2, Q.5, TP.HCM', 0, 1, 'C')
        pdf.cell(0, 10, 'ĐT: 0828.049.514 - 0123.456.789', 0, 1, 'C')
        pdf.set_font("NotoSans", style="B", size=12)
        pdf.cell(0, 15, 'BIÊN LAI KHÁCH HÀNG', 0, 1, 'C')
        pdf.set_font("NotoSans", style="", size=12)
        pdf.ln(5)
        pdf.set_left_margin(40)
        pdf.cell(100, 10, 'Ngày in: ' + str(date.today()), 'L')
        pdf.cell(100, 10, 'Số:' + data2Print[0], 'R')
        pdf.ln(5)
        pdf.cell(100, 10, 'Nhân viên: ' + data2Print[3], 'L', )
        pdf.cell(100, 10, 'In lúc:' + datetime.now().strftime("%H:%M:%S"), 'R')
        pdf.ln(5)
        pdf.cell(100, 10, 'Ngày ở: ' + str(dataPhieuThue[index][4]) , 'L')
        pdf.cell(100, 10, 'Ngày đi: ' + str(dataPhieuThue[index][5]) , 'R')
        pdf.ln(10)

        pdf.set_left_margin(40)
        for i in range(0, len(pttHeader)-3):
            if (i == 1):
                pdf.cell(28, 10, "Tên dịch vụ", border=1)
            elif (not i in [2, 3]):
                pdf.cell(28, 10, str(pttHeader[i]), border=1)
        pdf.ln(10)
        for i in range(0, len(data2Print)-3):
            if (i == 1):
                tempdv = ""
                for item in dataDv:
                    if (data2Print[i] == item[0]):
                        tempdv = item[1]
                pdf.cell(28, 10, tempdv, border=1)
            elif (i != 2 and i != 3):
                pdf.cell(28, 10, str(data2Print[i]), border=1)
        pdf.ln(10)

        pdf.set_left_margin(100)
        pdf.set_font("NotoSans", style="BI", size=12)
        pdf.cell(0, 10, 'Thuế VAT: ' + str(data2Print[6]), 'L')
        pdf.ln(10)
        pdf.cell(0, 10, 'Tổng tiền: ' + str(data2Print[7]), 'L')
        pdf.ln(10)
        
        pdf.output('bill.pdf', 'F')
       
    exportBotBut = Button(
        dashbrd,
        font = ('Chirp', 14),
        text = 'Export to PDF',
        bg = 'white',
        border = '1px solid black',
        height = 1,
        width = 15,
        command = printSelectedRow
    )

    sortBut2 = ttk.Combobox(
        dashbrd,
        font = ('Chirp', 10),
        values = ['Tăng dần', 'Giảm dần'],
        width = 20,
        cursor = 'hand2')
    sortBut2.set('--Thứ tự--')

    # ???

    srhBox1 = Text(
        height = 1,
        cursor = 'xterm',
        font = ('Chirp', 10),
        width = 20,
        border = '2px solid black'
    )

    # Tìm kiếm Nâng cao - Advanced Searching
    toRenderHeader1 = toRenderHeader
    global toRenderHeader2
    global toRenderHeader3
    global toRenderHeader4
    toRenderHeader2 = []
    toRenderHeader3 = []
    toRenderHeader4 = []
    sAdTextLbl1 = Label(dashbrd, text = 'Tìm kiếm nâng cao', bg = 'white')
    sAdTextLbl2 = Label(dashbrd, text = 'Từ', bg = 'white')
    sAdTextLbl3 = Label(dashbrd, text = 'Đến', bg = 'white')

    selected0 = StringVar()
    sortBut1 = ttk.Combobox(
        dashbrd,
        values = list(toRenderHeader1), 
        width = 20,
        textvariable = selected0
    )
    sortBut1.set('--Tiêu chí--')

    selected1 = StringVar()

    sAdLbl1 = ttk.Combobox(
        dashbrd,
        values = list(toRenderHeader1), 
        width = 20,
        textvariable= selected1
    )
    sAdLbl1.set('--Tiêu chí--')

    selected2 = StringVar()
    
    sAdLbl2 = ttk.Combobox(
        dashbrd,
        values = list(toRenderHeader2),
        width = 20,
        textvariable= selected2
    )
    sAdLbl2.set('--Giá trị--')

    sAdLbl3 = Text(
        dashbrd,
        font = ('Chirp', 10),
        width = 10,
        height = 1
    )

    sAdLbl4 = Text(
        dashbrd,
        font = ('Chirp', 10),
        width = 10,
        height = 1
    )
    
    def handleNormSearch():
        inp = srhBox1.get("1.0",'end-1c')
        # https://stackoverflow.com/questions/63525858/typeerror-get-missing-1-required-positional-argument-index1
        tmp = []
        for item in dataTblList4GenRight:
            for jtem in item:
                if (str(jtem).upper().find(inp.upper()) != -1):
                    tmp.append(item)
                    break
        global dataTblList
        dataTblList = tmp
        genRight(strTable, True)
        return
    def handleAdvancedSearch():
        # Link code: https://www.pythontutorial.net/tkinter/tkinter-combobox/
        tmp = []
        try:
            index = toRenderHeader1.index(selected1.get())
        except ValueError:
            messagebox.showwarning(title='Cảnh báo', message='Vui lòng chọn tiêu chí')
        # print(index)
        # print(dataTblList4GenRight[3])
        # print(dataTblList4GenRight[3][index])

        if (strTable == 'nhanVien'):
            stweird = 3
        else:
            stweird = 0
        if (not str(dataTblList4GenRight[stweird][index]).isnumeric()):
            messagebox.showwarning(title='Lỗi', message='Dữ liệu không phải số.')
            return
        x = sAdLbl3.get("1.0",'end-1c')
        y = sAdLbl4.get("1.0",'end-1c')
        if (x != '' and y != ''):
            if (x == ''): x = 0
            if (y == ''): y = 0
            x = int(x)
            y = int(y)
            if (x > y):
                messagebox.showwarning(title='Lỗi', message='Điểm dưới nhỏ hơn điểm trên.')
            else:
                tmp = []
                for row in dataTblList4GenRight:
                    if ((int(row[index]) > x) and (int(row[index]) < y) ):
                        # lần trước lỡ import lộn NoneType vào database
                        # https://stackoverflow.com/questions/3930188/how-to-convert-nonetype-to-int-or-string
                        tmp.append(row)
                global dataTblList
                oldDataHehe = dataTblList
                dataTblList = tmp
                genRight(strTable, True)
                dataTblList = oldDataHehe
        else:
            messagebox.showwarning(title='Lỗi', message='Bạn chưa nhập khoảng cần tìm.')
        return
    
    global icoSearch
    rawIcoSearch = Image.open('./img/sIco.png')
    rawIcoSearch = rawIcoSearch.resize((20, 20), Image.Resampling.LANCZOS)
    icoSearch = ImageTk.PhotoImage(rawIcoSearch)
    lblSrh1 = Button(dashbrd, image = icoSearch, font = ('Chirp', 14),bg = 'white', command = handleNormSearch)
    lblSrh2 = Button(dashbrd, image = icoSearch, font = ('Chirp', 14),bg = 'white', command = handleAdvancedSearch)

    # binding 
    def tmp1(e): 
        global toRenderHeader2
        toRenderHeader2 = []
        usrSel1 = e.widget.get()
        sAdLbl1.set(e.widget.get())
        if (tabSel[0] and usrSel1 in ptpHeader):
            for item in dataTblList4GenRight:
                toRenderHeader2.append(item[ptpHeader.index(usrSel1)])
            toRenderHeader2 = list(set(toRenderHeader2))
            sAdLbl2['values'] = toRenderHeader2
        if (tabSel[1] and usrSel1 in pttHeader):
            for item in dataTblList4GenRight:
                toRenderHeader2.append(item[pttHeader.index(usrSel1)])
            toRenderHeader2 = list(set(toRenderHeader2))
            sAdLbl2['values'] = toRenderHeader2
        if (tabSel[2] and usrSel1 in phongHeader):
            for item in dataTblList4GenRight:
                toRenderHeader2.append(item[phongHeader.index(usrSel1)])
            toRenderHeader2 = list(dict.fromkeys(toRenderHeader2))
            sAdLbl2['values'] = toRenderHeader2
        if (tabSel[3] and usrSel1 in nvHeader):
            for item in dataTblList4GenRight:
                toRenderHeader2.append(item[nvHeader.index(usrSel1)])
            toRenderHeader2 = list(set(toRenderHeader2))
            sAdLbl2['values'] = toRenderHeader2
        if (tabSel[4] and usrSel1 in khHeader):
            for item in dataTblList4GenRight:
                toRenderHeader2.append(item[khHeader.index(usrSel1)])
            toRenderHeader2 = list(set(toRenderHeader2))
            sAdLbl2['values'] = toRenderHeader2
        if (tabSel[5] and usrSel1 in dvHeader):
            for item in dataTblList4GenRight:
                toRenderHeader2.append(item[dvHeader.index(usrSel1)])
            toRenderHeader2 = list(set(toRenderHeader2))
            sAdLbl2['values'] = toRenderHeader2
        if (tabSel[5] and usrSel1 in pnHeader):
            for item in dataTblList4GenRight:
                toRenderHeader2.append(item[pnHeader.index(usrSel1)])
            toRenderHeader2 = list(set(toRenderHeader2))
            sAdLbl2['values'] = toRenderHeader2
    def tmp2(e): 
        sAdLbl2.set(e.widget.get())
        sample = e.widget.get()
        index = toRenderHeader1.index(selected1.get())
        tmp = []
        global dataTblList
        for row in dataTblList:
            if (sample == row[index]):
                tmp.append(row)
            elif (sample == str(row[index])):
                tmp.append(row)
        oldDataHehe = dataTblList
        dataTblList = tmp
        genRight(strTable, True)
        dataTblList = oldDataHehe
        return
    def tmp3(e):
        global dataTblList
        index = toRenderHeader1.index(selected0.get())
        if (e.widget.get() == 'Tăng dần'):
            dataTblList = sorted(dataTblList, key = itemgetter(index))
            genRight(strTable, True)
        else:
            dataTblList = sorted(dataTblList, key = itemgetter(index))
            dataTblList = dataTblList[::-1]
            genRight(strTable, True)
        return

    def parseAndSave(e):
        new_table = rSh.get_sheet_data(False, False, False)
        # print(strTable)
        cur = con.cursor()

        if (strTable == 'phieuThue'):
            try:
                cur.execute('drop table '+strTable)
            except sqlite3.OperationalError:
                pass
            cur.execute('''CREATE TABLE  '''
            + strTable
            + '''
                (id text, 
                idkh text,
                idnv text,
                iddv text,
                ngayden text,
                ngaydi text,
                tratruoc bool,
                thanhtien real,
                vat real,
                tongtien real,
                idphong text
            )''')
            for row in new_table:
                cur.execute(
                    '''
                    INSERT INTO {}
                    VALUES ('{}', '{}', '{}', '{}', '{}', '{}', {}, {}, {}, {}, '{}')
                    '''.format(strTable, row[0], row[1], row[2], row[3], row[4], row[5], str(row[6]), row[7], row[8], row[9], row[10])
                )
            con.commit()
            fetchFromDb()
        elif (strTable == 'phieuThanhToan'):
            try:
                cur.execute('drop table '+strTable)
            except sqlite3.OperationalError:
                pass
            cur.execute('''CREATE TABLE  '''
            + strTable
            + '''    
                (id text, 
                iddv text,
                idnv text,
                tennv text,
                songayo int,
                tongtien int,
                vat real,
                tienphaitra real,
                ngayin text
            )''')
            for row in new_table:
                cur.execute(
                    '''
                    INSERT INTO {}
                    VALUES ('{}', '{}', '{}', '{}', {}, {}, {}, {}, '{}')
                    '''.format(strTable, row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8])
                )
            con.commit()
        elif (strTable == 'phong'):
            try:
                cur.execute('drop table '+strTable)
            except sqlite3.OperationalError:
                pass
            cur.execute('''CREATE TABLE  '''
                + strTable
                + '''    
                    (id text, 
                    price int,
                    status text,
                    level text,
                    descript text,
                    items text
                )'''
            )
            for row in new_table:
                cur.execute('''
                    INSERT INTO {}
                    VALUES ('{}',{}, '{}', '{}', '{}', '{}')
                    '''.format(strTable, row[0], row[1], row[2], row[3], row[4], row[5])
                )
            con.commit()
        elif (strTable == 'nhanVien'):
            try:
                cur.execute('drop table '+strTable)
            except sqlite3.OperationalError:
                pass
            cur.execute('''CREATE TABLE  '''
            + strTable
            + '''    
                (id text, 
                ho text,
                ten text,
                gender text,
                year int,
                quequan text,
                chucvu text,
                tochuc text,
                luong int,
                thuong int
            )''')
            for row in new_table:
                cur.execute(
                    '''
                    INSERT INTO {}
                    VALUES ('{}', '{}', '{}', '{}', {}, '{}', '{}', '{}', {}, {})
                    '''.format(strTable, row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9])
                )
            con.commit()
        elif (strTable == 'khachHang'):
            try:
                cur.execute('drop table '+strTable)
            except sqlite3.OperationalError:
                pass
            cur.execute('''CREATE TABLE  '''
            + strTable
            + '''    
                (id text, 
                ho text,
                ten text,
                gender text,
                year int,
                diachi text,
                cmnd int,
                quoctich text,
                gmail text,
                sdt text
            )''')
            for row in new_table:
                cur.execute(
                    '''
                    INSERT INTO {}
                    VALUES ('{}', '{}', '{}', '{}', {}, '{}', {}, '{}', '{}', '{}')
                    '''.format(strTable, row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9])
                )
            con.commit()
            fetchFromDb()
        elif (strTable == 'dichVu'):
            try:
                cur.execute('drop table '+strTable)
            except sqlite3.OperationalError:
                pass
            cur.execute('''CREATE TABLE  '''
            + strTable
            + '''    
                (id text, 
                ten text,
                price int
            )''')
            for row in new_table:
                cur.execute(
                    '''
                    INSERT INTO {}
                    VALUES ('{}', '{}', {})
                    '''.format(strTable, row[0], row[1], row[2])
                )
            con.commit()
        elif (strTable == 'pNhapTbiAndFood'):
            try:
                cur.execute('drop table '+strTable)
            except sqlite3.OperationalError:
                pass
            cur.execute('''CREATE TABLE  '''
            + strTable
            + '''    
                (id text, 
                idfood text,
                ten text,
                sl int,
                idncc text,
                ncc text,
                idnv text,
                price int,
                date text
            )''')
            for row in new_table:
                cur.execute(
                    '''
                    INSERT INTO {}
                    VALUES ('{}', '{}', '{}', {}, '{}', '{}', '{}', {}, '{}')
                    '''.format(strTable, row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8])
                )
            con.commit()
            fetchFromDb()
        cur.execute('''
                    SELECT * FROM {}
                '''.format(strTable))
        # print(cur.fetchall())

        return

    sAdLbl1.bind("<<ComboboxSelected>>", tmp1)
    sAdLbl2.bind("<<ComboboxSelected>>", tmp2)
    sortBut2.bind("<<ComboboxSelected>>", tmp3)
    saveBotBut.bind("<Button-1>", parseAndSave)
    # end binding

    saveBotBut.place(x=775, y=630)
    exportXlsxBotBut.place(x = 500, y = 630)
    exportBotBut.place(x=900, y=630)

    srhBox1.place(x=390, y=110)
    lblSrh1.place(x=555, y=105)
    sortBut1.place(x=625, y=110)
    sortBut2.place(x=780, y=110)

    sAdTextLbl1.place(x=340, y=150)
    sAdTextLbl2.place(x=845, y=150)
    sAdTextLbl3.place(x=975, y=150)
    sAdLbl1.place(x=475, y =150)
    sAdLbl2.place(x=650, y =150)
    sAdLbl3.place(x=745+130, y =150)
    sAdLbl4.place(x=745+260, y =150)
    lblSrh2.place(x=1095, y= 150)







loggedIn = False

def getUsernameAndLogin(event=None):
    username = entryLogin.get() 
    password = entryPassword.get()
    passwordHashed = hashlib.sha256(password.encode()).hexdigest()
    with open('./json/userData.json', encoding='utf8') as userDataJsonFile:
        tempDict = json.load(userDataJsonFile)
    if (tempDict.get(username) == passwordHashed):
        global loggedIn
        loggedIn = True
        messagebox.showwarning(title=None, message = 'Login thành công. Thoát cửa sổ đăng nhập để hiển thị thanh điều khiển.')

    else:
        labelImgNotify.place(x=725, y=2.5)
        th_popUpWrongCredentialsLogin()
    return

def popUpWrongCredentialsLogin():
    time.sleep(2)
    try:
        labelImgNotify.place_forget()
    except Exception:
        print('nothing')
    return

def th_popUpWrongCredentialsLogin():
    thread = threading.Thread(target=popUpWrongCredentialsLogin)
    thread.start()
    return
class loginUI():
    def __init__(self, width, height, xaxis, yaxis):
        global loginFrame
        loginFrame = Tk()
        loginFrame.title('Đăng nhập cùng chúng tôi')
        loginFrame.iconbitmap('login.ico')
        loginFrame.configure(bg='white')
        self.width = width
        self.height = height
        self.xaxis = xaxis
        self.yaxis = yaxis
    def renderFrame(self):
        temp = str(self.width) + 'x' + str(self.height) + '+' + str(self.xaxis) + '+' +  str(self.yaxis)
        loginFrame.geometry(temp)

        imgIllus = PhotoImage(file = './img/undraw1.png')
        labelIllus = Label(loginFrame, image = imgIllus, bg='white')
        labelIllus.place(x=100, y=100)

        global dn_fr_img
        raw_dn_fr_img = Image.open('./img/dn.png')
        raw_dn_fr_img = raw_dn_fr_img.resize((int(544/1.4), int(629/1.4)), Image.Resampling.LANCZOS)
        dn_fr_img = ImageTk.PhotoImage(raw_dn_fr_img)
        lbl_dn_fr_img = Label(loginFrame, image = dn_fr_img)
        lbl_dn_fr_img.place(x = 560, y = 40);

        labelTitle = Label(
            loginFrame, 
            text = "ĐĂNG NHẬP", 
            font= ('Times New Roman', 24), 
            foreground= "blue3",
            bg='white'
        )
        labelTitle.place(x=675, y=90)

        emailPassEntryWidth = 30
        emailPassEntryHeight = 35

        labelLogin = Label(
            loginFrame, 
            text = "Email: ", 
            font= ('Times New Roman', 14), 
            foreground= "black",
            bg = 'white'
        )
        labelLogin.place(x=625, y=150)
        
        global entryLogin
        entryLogin = Entry(
            loginFrame, 
            borderwidth=2,
            font= ('Times New Roman', 14), 
            width=emailPassEntryWidth,

        )
        entryLogin.place(x=625, y=200, height=emailPassEntryHeight)
            
        labelPassword = Label(
            loginFrame, 
            text = "Mật khẩu: ", 
            font= ('Times New Roman', 14),
            foreground= "black",
            bg = 'white',
        )
        labelPassword.place(x=625, y=250)
        
        global entryPassword
        entryPassword = Entry(
            loginFrame, 
            show = '*', 
            borderwidth=2,
            font= ('Times New Roman', 14), 
            width=emailPassEntryWidth
        )
        entryPassword.place(x=625, y=300, height=emailPassEntryHeight)

        imgLoginBut = PhotoImage(file = './img/login.png')
        labelLoginBut = Label(loginFrame, image = imgLoginBut, bg='white')
        labelLoginBut.bind('<Button-1>', getUsernameAndLogin)
        bindedBut = Button(loginFrame, image=imgLoginBut, command=getUsernameAndLogin)
        labelLoginBut.place(x=665, y=370)

        # i don't know why it ran perfectly
        # Source: https://stackoverflow.com/questions/40658728/clickable-images-for-python
        
        global img
        global labelImgNotify
        img = PhotoImage(file = './img/vnwronglogin.png')
        labelImgNotify = Label(loginFrame, image = img, bg='white')

        

        loginFrame.mainloop()

myLoginUI = loginUI(1024, 576 , 150, 75)
myLoginUI.renderFrame()

if (loggedIn):
    dashbrd = Tk()
    genTopBanner()
    global icoAdd
    rawIcoAdd = Image.open('./img/add.png')
    rawIcoAdd = rawIcoAdd.resize((20, 20), Image.Resampling.LANCZOS)
    icoAdd = ImageTk.PhotoImage(rawIcoAdd)
    # startAutoSave()
    # startKeyListener()
    genDashUI()
    genNav()
    genRight('phieuThue', False)
    genBotBut('phieuThue')
    dashbrd.mainloop()
    winClosed = True
    # listener.join()

con.close()


